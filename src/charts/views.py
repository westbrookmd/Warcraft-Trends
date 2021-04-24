import openpyxl, json
from django.contrib.auth import get_user_model
from django.http import JsonResponse
from django.shortcuts import render
from django.views.generic import View
from rest_framework.views import APIView
from rest_framework.response import Response
#import json, requests
#import pandas as pd

#sheetName = '1JCVnNPNvjvecRMULcCNffGjbYFKb7Are0DvQcaS7G_A'
#ss = ezsheets.Spreadsheet(sheetName)

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

maxColumn = sheet.max_column


maxRow = sheet.max_row


produceNames = []
for i in range(2, 50):
    cellData = sheet.cell(row = i, column = 1)
    produceNames.append(cellData.value)


produceValues = []
for i in range(2, 50):
    cellData = sheet.cell(row = i, column = 2)
    produceValues.append(cellData.value)



with open('wow_issues.json') as json_file:
    wowData = json.load(json_file)

wowClasses = []
wowClassesUnclassified = 0
for issue in range(len(wowData)-1):
    try:
        wowClasses.append(wowData[issue]['labels'][-1]['name'])
    except:
        wowClassesUnclassified += 1



user = get_user_model()
Priests = sum('Priest' in s for s in wowClasses)
Hunters = sum('Hunter' in s for s in wowClasses)
Rogues = sum('Rogue' in s for s in wowClasses)
Warlocks = sum('Warlock' in s for s in wowClasses)
DeathKnights = sum('Death Knight' in s for s in wowClasses)
Mages = sum('Mage' in s for s in wowClasses)
Shamans = sum('Shaman' in s for s in wowClasses)
DemonHunters = sum('Demon' in s for s in wowClasses)
Druids = sum('Druid' in s for s in wowClasses)
Monks = sum('Monk' in s for s in wowClasses)
Paladins = sum('Paladin' in s for s in wowClasses)
Warriors = sum('Warrior' in s for s in wowClasses)

bug_count = (Priests + Hunters + Rogues +  Warlocks +  DeathKnights +  Mages + Shamans + DemonHunters + Druids + Monks + Paladins + Warriors)
#produce_count = len(produceNames)

class HomeView(View):
    def get(self, request, *args, **kwargs):
        return render(request, 'charts.html')



def get_data(request, *args, **kwargs):
    data = {
        "Total Bugs": bug_count,
        "Priests": Priests,
        "Hunters": Hunters,
        "Rogues": Rogues,
        "Warlocks": Warlocks,
        "DeathKnights": DeathKnights,
        "Mages": Mages,
        "Shamans": Shamans,
        "DemonHunters": DemonHunters,
        "Druids": Druids,
        "Monks": Monks,
        "Paladins": Paladins,
        "Warriors": Warriors,
    }
    return JsonResponse(data) # http response


class ChartData(APIView):
    authentication_classes = []
    permission_classes = []

    def get(self, request, format=None):
        labels = ["Priests", "Hunters", "Rogues", "Warlocks", "Death Knights", "Mages", "Shamans", "Demon Hunters", "Druids", "Monks", "Paladins", "Warriors"]
        default_items = [Priests, Hunters, Rogues, Warlocks, DeathKnights, Mages, Shamans, DemonHunters, Druids, Monks, Paladins, Warriors]
        data = {
                "labels": labels,
                "default": default_items,
        }
        return Response(data)

